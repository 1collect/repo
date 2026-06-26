from decimal import Decimal, InvalidOperation

from django.core.files import File
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    ImportBatch,
    OperationLog,
    Product,
    ProductHistory,
    Stock,
    StockHistory,
    Warehouse,
)


PRODUCT_SHEET = "\u043d\u043e\u043c\u0435\u043d\u043a\u043b\u0430\u0442\u0443\u0440\u0430"
WAREHOUSE_SHEET = "\u0441\u043a\u043b\u0430\u0434\u044b"
STOCK_SHEET = "\u043e\u0441\u0442\u0430\u0442\u043a\u0438"

H_NAME = "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
H_TITLE = "\u043d\u0430\u0437\u0432\u0430\u043d\u0438\u0435"
H_ARTICLE = "\u0430\u0440\u0442\u0438\u043a\u0443\u043b"
H_BARCODE = "\u0448\u0442\u0440\u0438\u0445\u043a\u043e\u0434"
H_CATEGORY = "\u043a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u044f"
H_DESCRIPTION = "\u043e\u043f\u0438\u0441\u0430\u043d\u0438\u0435"
H_WAREHOUSE = "\u0441\u043a\u043b\u0430\u0434"
H_WAREHOUSE_NAME = "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0441\u043a\u043b\u0430\u0434\u0430"
H_QUANTITY = "\u043a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e"
H_STOCK = "\u043e\u0441\u0442\u0430\u0442\u043e\u043a"


class ImportRejected(Exception):
    pass


class ImportHasErrors(Exception):
    pass


def _rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip().lower() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    result = []
    for number, values in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        result.append((number, dict(zip(headers, values))))
    return result


def _value(row, *names, required=True):
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    if required:
        raise ValueError(f"Missing required column: {names[0]}.")
    return ""


def _error(errors, sheet, row, message):
    errors.append({"sheet": sheet, "row": row, "message": str(message)})

def _preview_rows(rows_by_sheet, errors):
    errors_by_location = {}
    for item in errors:
        errors_by_location.setdefault((item["sheet"], item["row"]), []).append(item["message"])
    sheet_labels = {
        PRODUCT_SHEET: "Nomenclature",
        WAREHOUSE_SHEET: "Warehouses",
        STOCK_SHEET: "Receipts",
    }
    preview = []
    for sheet_key, rows in rows_by_sheet.items():
        label = sheet_labels.get(sheet_key, sheet_key)
        for row_number, row in rows:
            row_errors = errors_by_location.get((label, row_number), [])
            preview.append(
                {
                    "sheet": label,
                    "row": row_number,
                    "status": "error" if row_errors else "ok",
                    "errors": row_errors,
                    "values": {key: str(value) for key, value in row.items() if value is not None},
                }
            )
    return preview


def _finish_batch(batch, *, status, total_rows, applied_rows=0, errors=None, result=None, preview=None):
    errors = errors or []
    batch.status = status
    batch.total_rows = total_rows
    batch.applied_rows = applied_rows
    batch.error_rows = len(errors)
    batch.errors_json = errors or None
    batch.errors_text = "\n".join(
        f"{item['sheet']}, row {item['row']}: {item['message']}" for item in errors
    )
    batch.result_json = result or None
    batch.preview_json = preview or None
    batch.finished_at = timezone.now()
    batch.save(
        update_fields=[
            "status",
            "total_rows",
            "applied_rows",
            "error_rows",
            "errors_json",
            "errors_text",
            "result_json",
            "preview_json",
            "finished_at",
        ]
    )


def _product_snapshot(product):
    return {
        "name": product.name,
        "category": product.category,
        "article": product.article,
        "barcode": product.barcode,
        "description": product.description,
        "imported": product.imported,
        "is_active": product.is_active,
    }


def _operation_snapshot(user, *, warehouse=None, product=None):
    data = {
        "user_name_snapshot": str(user),
        "destination_name_snapshot": warehouse.name if warehouse else "",
    }
    if product:
        data.update(
            {
                "product_name_snapshot": product.name,
                "product_category_snapshot": product.category,
                "product_article_snapshot": product.article,
                "product_barcode_snapshot": product.barcode,
            }
        )
    return data


def _create_import_log(batch, user, action, *, status="", old_value=None, new_value=None, metadata=None):
    return OperationLog.objects.create(
        import_batch=batch,
        user=user,
        action=action,
        status=status,
        old_value=old_value,
        new_value=new_value,
        metadata=metadata,
        user_name_snapshot=str(user),
    )


def _existing_product_for(article, barcode):
    by_article = Product.objects.filter(article=article).first() if article else None
    by_barcode = Product.objects.filter(barcode=barcode).first() if barcode else None
    if by_article and by_barcode and by_article.pk != by_barcode.pk:
        raise ValueError(
            f"Article {article} belongs to {by_article}; barcode {barcode} belongs to {by_barcode}."
        )
    return by_article or by_barcode


def _parse_quantity(raw):
    try:
        quantity = Decimal(raw)
    except InvalidOperation as error:
        raise ValueError("Quantity must be an integer.") from error
    if quantity <= 0 or quantity != quantity.to_integral_value():
        raise ValueError("Receipt quantity must be an integer greater than zero.")
    return int(quantity)


def _collect_workbook(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = {name.lower(): name for name in workbook.sheetnames}
    missing = [name for name in [PRODUCT_SHEET, WAREHOUSE_SHEET, STOCK_SHEET] if name not in sheet_names]
    if missing:
        raise ValueError(f"Missing sheets: {', '.join(missing)}.")
    return {
        PRODUCT_SHEET: _rows(workbook[sheet_names[PRODUCT_SHEET]]),
        WAREHOUSE_SHEET: _rows(workbook[sheet_names[WAREHOUSE_SHEET]]),
        STOCK_SHEET: _rows(workbook[sheet_names[STOCK_SHEET]]),
    }


def _validate_rows(rows_by_sheet):
    errors = []
    product_rows = []
    warehouse_rows = []
    stock_rows = []
    seen_articles = {}
    seen_barcodes = {}
    product_articles = set()
    product_barcodes = set()
    warehouse_names = set()

    for row_number, row in rows_by_sheet[PRODUCT_SHEET]:
        try:
            name = _value(row, H_NAME, H_TITLE)
            article = _value(row, H_ARTICLE)
            barcode = _value(row, H_BARCODE)
            category = _value(row, H_CATEGORY, required=False)
            description = _value(row, H_DESCRIPTION, required=False)
            if len(article) != 5 or not article.isdigit():
                raise ValueError("Article must contain exactly 5 digits.")
            if article in seen_articles:
                raise ValueError(f"Duplicate article {article} in row {seen_articles[article]}.")
            if barcode in seen_barcodes:
                raise ValueError(f"Duplicate barcode {barcode} in row {seen_barcodes[barcode]}.")
            seen_articles[article] = row_number
            seen_barcodes[barcode] = row_number
            product = _existing_product_for(article, barcode)
            name_owner = Product.objects.filter(name__iexact=name).exclude(
                Q(article=article) | Q(barcode=barcode)
            ).first()
            if name_owner:
                raise ValueError(f"Product name already belongs to article {name_owner.article}.")
            product_rows.append(
                {
                    "row": row_number,
                    "name": name,
                    "article": article,
                    "barcode": barcode,
                    "category": category,
                    "description": description,
                    "product": product,
                }
            )
            product_articles.add(article)
            product_barcodes.add(barcode)
        except Exception as error:
            _error(errors, "Nomenclature", row_number, error)

    for row_number, row in rows_by_sheet[WAREHOUSE_SHEET]:
        try:
            name = _value(row, H_NAME, H_TITLE, H_WAREHOUSE)
            if name in warehouse_names:
                raise ValueError(f"Duplicate warehouse {name}.")
            warehouse_names.add(name)
            warehouse_rows.append({"row": row_number, "name": name})
        except Exception as error:
            _error(errors, "Warehouses", row_number, error)

    for row_number, row in rows_by_sheet[STOCK_SHEET]:
        try:
            warehouse_name = _value(row, H_WAREHOUSE, H_WAREHOUSE_NAME)
            article = _value(row, H_ARTICLE, required=False)
            barcode = _value(row, H_BARCODE, required=False)
            if not article and not barcode:
                raise ValueError("Provide article or barcode.")
            quantity = _parse_quantity(_value(row, H_QUANTITY, H_STOCK))
            if warehouse_name not in warehouse_names and not Warehouse.objects.filter(name=warehouse_name).exists():
                raise ValueError(f"Warehouse {warehouse_name} was not found.")
            product = _existing_product_for(article, barcode)
            if not product and article not in product_articles and barcode not in product_barcodes:
                raise ValueError(f"Product {article or barcode} was not found.")
            stock_rows.append(
                {
                    "row": row_number,
                    "warehouse_name": warehouse_name,
                    "article": article,
                    "barcode": barcode,
                    "quantity": quantity,
                }
            )
        except Exception as error:
            _error(errors, "Receipts", row_number, error)

    return product_rows, warehouse_rows, stock_rows, errors


def _apply_import(batch, user, product_rows, warehouse_rows, stock_rows):
    result = {"products": 0, "warehouses": 0, "stocks": 0}
    product_by_article = {}
    product_by_barcode = {}

    with transaction.atomic():
        for item in warehouse_rows:
            warehouse, created = Warehouse.objects.get_or_create(
                name=item["name"],
                defaults={"is_active": False},
            )
            if created:
                OperationLog.objects.create(
                    import_batch=batch,
                    user=user,
                    action="Excel warehouse import",
                    destination=warehouse,
                    status="Warehouse created",
                    new_value={"name": warehouse.name, "is_active": warehouse.is_active},
                    **_operation_snapshot(user, warehouse=warehouse),
                )
            result["warehouses"] += 1

        for item in product_rows:
            product = _existing_product_for(item["article"], item["barcode"])
            data = {
                "name": item["name"],
                "category": item["category"],
                "article": item["article"],
                "barcode": item["barcode"],
                "description": item["description"],
                "imported": True,
            }
            if product:
                old_value = _product_snapshot(product)
                changed_fields = []
                for field, value in data.items():
                    if getattr(product, field) != value:
                        setattr(product, field, value)
                        changed_fields.append(field)
                if changed_fields:
                    product.save(update_fields=changed_fields)
                    new_value = _product_snapshot(product)
                    ProductHistory.objects.create(
                        product=product,
                        changed_by=user,
                        import_batch=batch,
                        old_value=old_value,
                        new_value=new_value,
                        changed_fields=changed_fields,
                    )
                    OperationLog.objects.create(
                        import_batch=batch,
                        user=user,
                        action="Excel product update",
                        product=product,
                        status="Product updated",
                        old_value=old_value,
                        new_value=new_value,
                        **_operation_snapshot(user, product=product),
                    )
            else:
                product = Product.objects.create(**data)
                OperationLog.objects.create(
                    import_batch=batch,
                    user=user,
                    action="Excel product import",
                    product=product,
                    status="Product created",
                    new_value=_product_snapshot(product),
                    **_operation_snapshot(user, product=product),
                )
            product_by_article[product.article] = product
            product_by_barcode[product.barcode] = product
            result["products"] += 1

        for item in stock_rows:
            warehouse = Warehouse.objects.get(name=item["warehouse_name"])
            product = (
                product_by_article.get(item["article"])
                or product_by_barcode.get(item["barcode"])
                or _existing_product_for(item["article"], item["barcode"])
            )
            stock, _ = Stock.objects.select_for_update().get_or_create(
                warehouse=warehouse,
                product=product,
                defaults={"quantity": 0},
            )
            old_quantity = stock.quantity
            new_quantity = old_quantity + item["quantity"]
            stock.quantity = new_quantity
            stock.save(update_fields=["quantity", "updated_at"])
            StockHistory.objects.create(
                stock=stock,
                warehouse=warehouse,
                product=product,
                changed_by=user,
                import_batch=batch,
                old_quantity=old_quantity,
                change_quantity=item["quantity"],
                new_quantity=new_quantity,
                reason="Excel import receipt",
            )
            OperationLog.objects.create(
                import_batch=batch,
                user=user,
                action="Excel receipt",
                destination=warehouse,
                product=product,
                quantity=item["quantity"],
                resulting_quantity=new_quantity,
                status="Stock increased",
                old_value={"quantity": old_quantity},
                new_value={"quantity": new_quantity},
                **_operation_snapshot(user, warehouse=warehouse, product=product),
            )
            result["stocks"] += 1

        _create_import_log(
            batch,
            user,
            "Excel import succeeded",
            status="Import applied",
            new_value=result,
        )
    return result


def prepare_iiko_workbook(file, user):
    original_filename = getattr(file, "name", None) or "import.xlsx"
    upload = file if getattr(file, "name", None) else File(file, name=original_filename)
    batch = ImportBatch.objects.create(
        uploaded_by=user,
        original_filename=original_filename,
        file=upload,
    )
    total_rows = 0
    try:
        rows_by_sheet = _collect_workbook(batch.file.path)
        total_rows = sum(len(rows) for rows in rows_by_sheet.values())
        product_rows, warehouse_rows, stock_rows, errors = _validate_rows(rows_by_sheet)
        preview = _preview_rows(rows_by_sheet, errors)
        if errors:
            _finish_batch(
                batch,
                status=ImportBatch.Status.FAILED,
                total_rows=total_rows,
                errors=errors,
                preview=preview,
            )
            _create_import_log(
                batch,
                user,
                "Excel import preview failed",
                status="Preview has errors",
                old_value={"file": batch.original_filename},
                new_value={"errors": errors},
            )
        else:
            _finish_batch(
                batch,
                status=ImportBatch.Status.PENDING,
                total_rows=total_rows,
                preview=preview,
            )
            _create_import_log(
                batch,
                user,
                "Excel import preview ready",
                status="Ready to apply",
                old_value={"file": batch.original_filename},
                new_value={"rows": total_rows},
            )
        return {"success": not bool(errors), "batch": batch, "errors": errors}
    except Exception as error:
        errors = [{"sheet": "File", "row": "-", "message": str(error)}]
        _finish_batch(
            batch,
            status=ImportBatch.Status.FAILED,
            total_rows=total_rows,
            errors=errors,
            preview=locals().get("preview"),
        )
        _create_import_log(
            batch,
            user,
            "Excel import preview failed",
            status="Preview rejected",
            old_value={"file": batch.original_filename},
            new_value={"errors": errors},
        )
        return {"success": False, "batch": batch, "errors": errors}


def apply_iiko_import_batch(batch, user):
    if batch.error_rows:
        raise ImportHasErrors("Сначала исправьте ошибки в Excel и загрузите файл повторно")
    if batch.status == ImportBatch.Status.SUCCESS:
        return batch.result_json or {"products": 0, "warehouses": 0, "stocks": 0}
    rows_by_sheet = _collect_workbook(batch.file.path)
    product_rows, warehouse_rows, stock_rows, errors = _validate_rows(rows_by_sheet)
    if errors:
        preview = _preview_rows(rows_by_sheet, errors)
        _finish_batch(
            batch,
            status=ImportBatch.Status.FAILED,
            total_rows=sum(len(rows) for rows in rows_by_sheet.values()),
            errors=errors,
            preview=preview,
        )
        raise ImportHasErrors("Сначала исправьте ошибки в Excel и загрузите файл повторно")
    result = _apply_import(batch, user, product_rows, warehouse_rows, stock_rows)
    _finish_batch(
        batch,
        status=ImportBatch.Status.SUCCESS,
        total_rows=sum(len(rows) for rows in rows_by_sheet.values()),
        applied_rows=len(product_rows) + len(warehouse_rows) + len(stock_rows),
        result=result,
        preview=_preview_rows(rows_by_sheet, []),
    )
    return result


def import_iiko_workbook(file, user):
    result = prepare_iiko_workbook(file, user)
    batch = result["batch"]
    if not result.get("success"):
        return result
    totals = apply_iiko_import_batch(batch, user)
    return totals