from io import BytesIO
from unittest.mock import patch
from django.core.exceptions import PermissionDenied, ValidationError
from django.test import TestCase, override_settings
from django.utils import timezone
from django.urls import reverse
from openpyxl import Workbook
from .forms import ManagementUserForm
from .importers import import_iiko_workbook
from .models import OperationLog, Product, Stock, Transfer, User, Warehouse
from .notifications import _recipient_and_buttons
from .management.commands.runreminders import send_due_reminders
from .transfer_scan_service import (
    confirm_receive,
    confirm_shipment,
    find_active_transfers_by_barcode,
)
from .services import (
    approve_transfer,
    generate_iiko_barcode,
    create_transfer,
    reject_transfer,
    receive_transfer,
    receive_stock,
    sender_approve_transfer,
    sender_reject_transfer,
    ship_transfer,
)


class TransferFlowTests(TestCase):
    def setUp(self):
        self.sender = User.objects.create_user("sender", password="test-pass")
        self.receiver = User.objects.create_user("receiver", password="test-pass")
        self.admin_reviewer = User.objects.create_user(
            "admin-reviewer", password="test-pass", role=User.Role.ADMIN
        )
        self.source = Warehouse.objects.create(
            name="Алматы",
            responsible=self.sender,
        )
        self.destination = Warehouse.objects.create(name="Астана", responsible=self.receiver)
        self.product = Product.objects.create(
            name="Стул", article="12345", barcode="21006830064"
        )
        Stock.objects.create(warehouse=self.source, product=self.product, quantity=10)

    def test_every_transfer_requires_approval(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=3,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        self.assertEqual(transfer.status, Transfer.Status.WAITING_SENDER_APPROVAL)

        sender_approve_transfer(transfer.pk, self.sender)
        approve_transfer(transfer.pk, self.admin_reviewer)
        ship_transfer(transfer.pk, self.sender)
        self.assertEqual(
            Stock.objects.get(warehouse=self.source, product=self.product).quantity, 7
        )

        receive_transfer(transfer.pk, self.receiver)
        self.assertEqual(
            Stock.objects.get(warehouse=self.destination, product=self.product).quantity, 3
        )
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.COMPLETED)
        self.assertEqual(transfer.operations.count(), 5)

    def test_flow_with_approval(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=2,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        self.assertEqual(transfer.status, Transfer.Status.WAITING_SENDER_APPROVAL)
        sender_approve_transfer(transfer.pk, self.sender)
        approve_transfer(transfer.pk, self.admin_reviewer)
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.APPROVED)

    def test_cannot_ship_twice(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=3,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        sender_approve_transfer(transfer.pk, self.sender)
        approve_transfer(transfer.pk, self.admin_reviewer)
        ship_transfer(transfer.pk, self.sender)
        with self.assertRaises(ValidationError):
            ship_transfer(transfer.pk, self.sender)
        self.assertEqual(
            Stock.objects.get(warehouse=self.source, product=self.product).quantity, 7
        )

    def test_cannot_create_more_than_stock(self):
        with self.assertRaises(ValidationError):
            create_transfer(
                source=self.source,
                destination=self.destination,
                product=self.product,
                quantity=11,
                user=self.receiver,
                reason="Новый сотрудник",
            )

    def test_wrong_user_cannot_receive(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        sender_approve_transfer(transfer.pk, self.sender)
        approve_transfer(transfer.pk, self.admin_reviewer)
        ship_transfer(transfer.pk, self.sender)
        with self.assertRaises(PermissionDenied):
            receive_transfer(transfer.pk, self.sender)

    def test_main_pages_render(self):
        self.client.force_login(self.sender)
        for url in ["/", "/products/", "/stocks/", "/transfers/", "/operations/"]:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200, url)

    def test_dashboard_shows_only_latest_operation_per_product(self):
        first = OperationLog.objects.create(
            user=self.sender,
            action="Старая операция",
            product=self.product,
            product_name_snapshot=self.product.name,
            product_article_snapshot=self.product.article,
            user_name_snapshot=str(self.sender),
        )
        second = OperationLog.objects.create(
            user=self.sender,
            action="Новая операция",
            product=self.product,
            product_name_snapshot=self.product.name,
            product_article_snapshot=self.product.article,
            user_name_snapshot=str(self.sender),
        )
        self.client.force_login(self.sender)
        response = self.client.get(reverse("dashboard"))
        operations = response.context["latest_operations"]
        self.assertIn(second, operations)
        self.assertNotIn(first, operations)

    def test_transfer_api(self):
        self.client.force_login(self.receiver)
        response = self.client.post(
            "/api/transfers/",
            {
                "source": self.source.pk,
                "destination": self.destination.pk,
                "product": self.product.pk,
                "quantity": 2,
                "reason": "Новый сотрудник",
            },
        )
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()["status"], Transfer.Status.WAITING_SENDER_APPROVAL)

    def test_transfer_does_not_require_assigned_admin_reviewer(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        self.assertEqual(transfer.status, Transfer.Status.WAITING_SENDER_APPROVAL)

    def test_transfer_keeps_historical_names_and_codes(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=2,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        self.source.name = "Новое название склада"
        self.source.save(update_fields=["name"])
        self.product.name = "Новое название товара"
        self.product.article = "54321"
        self.product.barcode = "2100000000999"
        self.product.save(update_fields=["name", "article", "barcode"])

        transfer.refresh_from_db()
        self.assertEqual(transfer.source_name_snapshot, "Алматы")
        self.assertEqual(transfer.product_name_snapshot, "Стул")
        self.assertEqual(transfer.product_article_snapshot, "12345")
        self.assertEqual(transfer.product_barcode_snapshot, "21006830064")
        operation = transfer.operations.get(action="Создано перемещение")
        self.assertEqual(operation.source_name_snapshot, "Алматы")
        self.assertEqual(operation.product_name_snapshot, "Стул")
        self.assertEqual(operation.transfer_reason, "Новый сотрудник")

    def test_sender_cannot_create_request_to_other_warehouse(self):
        with self.assertRaises(PermissionDenied):
            create_transfer(
                source=self.source,
                destination=self.destination,
                product=self.product,
                quantity=1,
                user=self.sender,
                reason="Новый сотрудник",
            )

    def test_reject_requires_reason_and_saves_decision_details(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Новый сотрудник",
        )
        with self.assertRaises(ValidationError):
            reject_transfer(transfer.pk, self.admin_reviewer)

        sender_approve_transfer(transfer.pk, self.sender)
        with self.assertRaises(ValidationError):
            reject_transfer(transfer.pk, self.admin_reviewer)

        reject_transfer(transfer.pk, self.admin_reviewer, "Недостаточно обоснования")
        transfer.refresh_from_db()
        self.assertEqual(transfer.status, Transfer.Status.ADMIN_REJECTED)
        self.assertEqual(transfer.rejection_reason, "Недостаточно обоснования")
        self.assertEqual(transfer.rejected_by, self.admin_reviewer)
        operation = transfer.operations.get(action="Перемещение отклонено")
        self.assertEqual(operation.admin_decision, "Отклонено")
        self.assertEqual(operation.rejection_reason, "Недостаточно обоснования")

    def test_sender_can_reject_with_reason(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Новый сотрудник",
        )

        sender_reject_transfer(transfer.pk, self.sender, "Нет свободного остатка под заявку")
        transfer.refresh_from_db()

        self.assertEqual(transfer.status, Transfer.Status.SENDER_REJECTED)
        self.assertEqual(transfer.sender_rejection_reason, "Нет свободного остатка под заявку")
        self.assertEqual(transfer.sender_rejected_by, self.sender)


class ExcelImportTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="test-pass")

    def workbook(self):
        book = Workbook()
        products = book.active
        products.title = "Номенклатура"
        products.append(["Наименование", "Артикул", "Штрихкод", "Категория"])
        products.append(["Стул офисный", "12345", "21006830064", "Мебель"])
        warehouses = book.create_sheet("Склады")
        warehouses.append(["Наименование"])
        warehouses.append(["Склад Алматы"])
        stocks = book.create_sheet("Остатки")
        stocks.append(["Склад", "Артикул", "Количество"])
        stocks.append(["Склад Алматы", "12345", 10])
        content = BytesIO()
        book.save(content)
        content.seek(0)
        return content

    def test_import_preserves_codes_and_stock(self):
        result = import_iiko_workbook(self.workbook(), self.admin)
        product = Product.objects.get(article="12345")
        self.assertEqual(product.barcode, "21006830064")
        self.assertTrue(product.imported)
        self.assertEqual(Stock.objects.get(product=product).quantity, 10)
        self.assertEqual(result, {"products": 1, "warehouses": 1, "stocks": 1})


@override_settings(TELEGRAM_BOT_TOKEN="")
class ManagementTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            "manager",
            password="test-pass",
            role=User.Role.ADMIN,
        )
        self.regular_user = User.objects.create_user("worker", password="test-pass")

    def test_only_main_admin_can_open_management(self):
        self.client.force_login(self.regular_user)
        self.assertEqual(self.client.get(reverse("management")).status_code, 403)

        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse("management")).status_code, 200)

    def test_admin_can_create_user_with_hashed_password(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("management_user_create"),
            {
                "username": "worker2",
                "first_name": "Айжан",
                "last_name": "",
                "email": "",
                "role": User.Role.RESPONSIBLE,
                "is_active": "on",
                "password": "secure-pass",
            },
        )
        self.assertRedirects(response, reverse("management"))
        user = User.objects.get(username="worker2")
        self.assertTrue(user.check_password("secure-pass"))
        self.assertEqual(user.role, User.Role.RESPONSIBLE)

    def test_user_management_has_only_admin_and_user_roles(self):
        form = ManagementUserForm(request_user=self.admin)
        self.assertEqual(
            list(form.fields["role"].choices),
            [
                (User.Role.ADMIN, "Админ"),
                (User.Role.RESPONSIBLE, "Пользователь"),
            ],
        )

    def test_duplicate_username_shows_form_error(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("management_user_create"),
            {
                "username": self.regular_user.username,
                "first_name": "Другой",
                "last_name": "",
                "email": "",
                "role": User.Role.RESPONSIBLE,
                "is_active": "on",
                "password": "secure-pass",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Пользователь с таким именем уже существует")

    def test_admin_can_create_warehouse_assignments(self):
        responsible = User.objects.create_user(
            "responsible",
            password="test-pass",
            role=User.Role.RESPONSIBLE,
        )
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("management_warehouse_create"),
            {
                "name": "Центральный склад",
                "responsible": responsible.pk,
                "is_active": "on",
            },
        )
        self.assertRedirects(response, reverse("management"))
        warehouse = Warehouse.objects.get(name="Центральный склад")
        self.assertEqual(warehouse.responsible, responsible)

    def test_admin_can_unlink_user_telegram(self):
        self.regular_user.telegram_id = 123456789
        self.regular_user.save(update_fields=["telegram_id"])
        old_token = self.regular_user.telegram_link_token

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("management_user_unlink_telegram", args=[self.regular_user.pk])
        )

        self.assertRedirects(response, reverse("management"))
        self.regular_user.refresh_from_db()
        self.assertIsNone(self.regular_user.telegram_id)
        self.assertNotEqual(self.regular_user.telegram_link_token, old_token)

    def test_regular_user_cannot_unlink_telegram(self):
        self.client.force_login(self.regular_user)
        response = self.client.post(
            reverse("management_user_unlink_telegram", args=[self.admin.pk])
        )
        self.assertEqual(response.status_code, 403)


class TelegramNotificationTests(TestCase):
    def setUp(self):
        self.sender = User.objects.create_user("tg-sender", password="test-pass")
        self.receiver = User.objects.create_user("tg-receiver", password="test-pass")
        self.admin_reviewer = User.objects.create_user(
            "tg-admin-reviewer",
            password="test-pass",
            role=User.Role.ADMIN,
        )
        self.source = Warehouse.objects.create(
            name="Telegram отправитель",
            responsible=self.sender,
        )
        self.destination = Warehouse.objects.create(
            name="Telegram получатель",
            responsible=self.receiver,
        )
        self.product = Product.objects.create(
            name="Telegram товар",
            article="67890",
            barcode="2100000000091",
        )
        Stock.objects.create(warehouse=self.source, product=self.product, quantity=5)

    def test_waiting_sender_approval_has_sender_button(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Недостаток оборудования",
        )

        recipient, buttons = _recipient_and_buttons(transfer)

        self.assertEqual(recipient, self.sender)
        self.assertEqual(
            buttons,
            [
                ("Подтвердить согласие", f"transfer:sender_approve:{transfer.pk}"),
            ],
        )

    def test_reject_action_finishes_approval_flow(self):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Недостаток оборудования",
        )

        sender_approve_transfer(transfer.pk, self.sender)
        reject_transfer(transfer.pk, self.admin_reviewer, "Заявка создана ошибочно")
        transfer.refresh_from_db()

        self.assertEqual(transfer.status, Transfer.Status.ADMIN_REJECTED)
        recipient, buttons = _recipient_and_buttons(transfer)
        self.assertIsNone(recipient)
        self.assertEqual(buttons, [])


class ProductLabelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("label-user", password="test-pass")
        self.product = Product.objects.create(
            name="Монитор Vega",
            article="00018",
            barcode="2100683006415",
        )

    def test_label_requires_login(self):
        response = self.client.get(reverse("product_label", args=[self.product.pk]))
        self.assertEqual(response.status_code, 302)

    def test_label_contains_product_and_barcode_svg(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("product_label", args=[self.product.pk]),
            {"quantity": 3},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Монитор Vega")
        self.assertContains(response, "00018")
        self.assertContains(response, "КОЛ-ВО: 3")
        self.assertContains(response, "2100683006415")
        self.assertContains(response, "data:image/svg+xml;base64,")


class StockReceiptTests(TestCase):
    def setUp(self):
        self.responsible = User.objects.create_user(
            "stock-manager",
            password="test-pass",
            role=User.Role.RESPONSIBLE,
        )
        self.other_user = User.objects.create_user("other", password="test-pass")
        self.warehouse = Warehouse.objects.create(
            name="Склад техники",
            responsible=self.responsible,
        )
        self.product = Product.objects.create(
            name="Холодильник Brand X",
            article="54321",
            barcode="2100683006415",
        )

    def test_receipt_increases_stock_and_creates_history(self):
        stock = receive_stock(
            warehouse=self.warehouse,
            product=self.product,
            quantity=5,
            user=self.responsible,
        )
        self.assertEqual(stock.quantity, 5)
        operation = OperationLog.objects.get(action="Поступление товара")
        self.assertIsNone(operation.transfer)
        self.assertEqual(operation.destination, self.warehouse)
        self.assertEqual(operation.product, self.product)
        self.assertEqual(operation.quantity, 5)

    def test_second_receipt_adds_to_existing_stock(self):
        Stock.objects.create(warehouse=self.warehouse, product=self.product, quantity=2)
        stock = receive_stock(
            warehouse=self.warehouse,
            product=self.product,
            quantity=5,
            user=self.responsible,
        )
        self.assertEqual(stock.quantity, 7)

    def test_other_user_cannot_receive_to_warehouse(self):
        with self.assertRaises(PermissionDenied):
            receive_stock(
                warehouse=self.warehouse,
                product=self.product,
                quantity=5,
                user=self.other_user,
            )

    def test_receipt_view_and_history(self):
        self.client.force_login(self.responsible)
        response = self.client.post(
            reverse("stock_receipt"),
            {
                "warehouse": self.warehouse.pk,
                "product": self.product.pk,
                "quantity": 5,
            },
        )
        self.assertRedirects(response, reverse("stock_list"))
        history = self.client.get(reverse("operation_list"))
        self.assertContains(history, "Поступление товара")
        self.assertContains(history, "Холодильник Brand X")
        self.assertContains(history, "Склад техники")

    def test_responsible_cannot_create_new_product_from_receipt(self):
        self.client.force_login(self.responsible)
        response = self.client.get(reverse("stock_receipt"))
        self.assertNotContains(response, "Новая номенклатура")

    def test_admin_can_create_product_during_receipt(self):
        admin = User.objects.create_user(
            "receipt-admin",
            password="test-pass",
            role=User.Role.ADMIN,
        )
        self.client.force_login(admin)
        response = self.client.post(
            reverse("stock_receipt"),
            {
                "warehouse": self.warehouse.pk,
                "product": "",
                "quantity": 4,
                "new_product_name": "Монитор LG 24MP400",
                "new_product_category": "Мониторы",
                "new_product_description": "",
            },
        )
        self.assertRedirects(response, reverse("stock_list"))
        product = Product.objects.get(name="Монитор LG 24MP400")
        self.assertEqual(
            Stock.objects.get(warehouse=self.warehouse, product=product).quantity,
            4,
        )


@override_settings(
    TRANSFER_REMINDER_AFTER_HOURS=1,
    TRANSFER_REMINDER_REPEAT_HOURS=24,
)
class TransferReminderTests(TestCase):
    def setUp(self):
        self.sender = User.objects.create_user("reminder-sender", password="test-pass")
        self.receiver = User.objects.create_user("reminder-receiver", password="test-pass")
        self.admin_reviewer = User.objects.create_user(
            "reminder-admin-reviewer",
            password="test-pass",
            role=User.Role.ADMIN,
            telegram_id=111,
        )
        self.source = Warehouse.objects.create(
            name="Склад напоминаний",
            responsible=self.sender,
        )
        self.destination = Warehouse.objects.create(
            name="Получатель напоминаний",
            responsible=self.receiver,
        )
        self.product = Product.objects.create(
            name="Товар напоминаний",
            article="98765",
            barcode="2100000000982",
        )
        Stock.objects.create(warehouse=self.source, product=self.product, quantity=2)

    @patch("core.management.commands.runreminders.notify_transfer_reminder")
    def test_due_transfer_gets_one_reminder_and_timestamp(self, notify):
        transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=1,
            user=self.receiver,
            reason="Временное использование",
        )
        Transfer.objects.filter(pk=transfer.pk).update(
            created_at=timezone.now() - timezone.timedelta(hours=2)
        )

        self.assertEqual(send_due_reminders(), 1)
        notify.assert_called_once()
        transfer.refresh_from_db()
        self.assertIsNotNone(transfer.last_reminded_at)

        self.assertEqual(send_due_reminders(), 0)


class ProductCreationTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            "catalog-admin",
            password="test-pass",
            role=User.Role.ADMIN,
        )
        self.warehouse = Warehouse.objects.create(
            name="Главный склад",
            responsible=self.admin,
        )

    def test_creation_requires_category_warehouse_and_quantity(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("product_create"),
            {
                "name": "Холодильник Vega",
                "category": "",
                "description": "",
                "warehouse": self.warehouse.pk,
                "quantity": 5,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFormError(response.context["form"], "category", "Обязательное поле.")

    def test_creation_adds_initial_stock_and_history(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("product_create"),
            {
                "name": "Холодильник Vega",
                "category": "Холодильники",
                "description": "Модель X200",
                "warehouse": self.warehouse.pk,
                "quantity": 5,
            },
        )
        self.assertRedirects(response, reverse("product_list"))
        product = Product.objects.get(name="Холодильник Vega")
        self.assertEqual(product.category, "Холодильники")
        self.assertEqual(product.barcode, generate_iiko_barcode(product.article))
        self.assertEqual(
            Stock.objects.get(warehouse=self.warehouse, product=product).quantity,
            5,
        )
        operation = OperationLog.objects.get(
            action="Первичное поступление",
            product=product,
        )
        self.assertEqual(operation.quantity, 5)
        self.assertEqual(operation.destination, self.warehouse)

    def test_product_category_filter(self):
        Product.objects.create(
            name="Холодильник",
            category="Холодильники",
            article="10001",
            barcode="2100000000015",
        )
        Product.objects.create(
            name="Шкаф",
            category="Мебель",
            article="10002",
            barcode="2100000000022",
        )
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("product_list"),
            {"category": "Холодильники"},
        )
        self.assertContains(response, "Холодильник")
        self.assertNotContains(response, "Шкаф")

    def test_iiko_compatible_barcode_generation(self):
        self.assertEqual(generate_iiko_barcode("00010"), "2100010000019")
        self.assertEqual(generate_iiko_barcode("00014"), "2100014000015")
        self.assertEqual(generate_iiko_barcode("00018"), "2100018000011")
        self.assertEqual(generate_iiko_barcode("00409"), "2100409000019")

    def test_article_generation_uses_first_free_value(self):
        Product.objects.create(
            name="Первый товар",
            category="Техника",
            article="00001",
            barcode="2100001000018",
        )
        Product.objects.create(
            name="Второй товар",
            category="Техника",
            article="00002",
            barcode="2100002000017",
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("product_create"),
            {
                "name": "Телефон Grandstream",
                "category": "Техника",
                "description": "",
                "warehouse": self.warehouse.pk,
                "quantity": 1,
            },
        )

        self.assertRedirects(response, reverse("product_list"))
        product = Product.objects.get(name="Телефон Grandstream")
        self.assertEqual(product.article, "00003")
        self.assertEqual(product.barcode, generate_iiko_barcode("00003"))

class TransferScanServiceTests(TestCase):
    def setUp(self):
        self.sender = User.objects.create_user("scan-sender", password="test-pass")
        self.receiver = User.objects.create_user("scan-receiver", password="test-pass")
        self.other = User.objects.create_user("scan-other", password="test-pass")
        self.admin = User.objects.create_user("scan-admin", password="test-pass", role=User.Role.ADMIN)
        self.source = Warehouse.objects.create(name="Scan Алматы", responsible=self.sender)
        self.destination = Warehouse.objects.create(name="Scan Астана", responsible=self.receiver)
        self.product = Product.objects.create(name="Scan товар", article="54321", barcode="2105432100018")
        Stock.objects.create(warehouse=self.source, product=self.product, quantity=5)
        self.transfer = create_transfer(
            source=self.source,
            destination=self.destination,
            product=self.product,
            quantity=2,
            user=self.receiver,
            reason="Scan test",
        )
        sender_approve_transfer(self.transfer.pk, self.sender)
        approve_transfer(self.transfer.pk, self.admin)

    def test_sender_finds_approved_transfer_by_barcode(self):
        transfers = find_active_transfers_by_barcode(self.sender, self.product.barcode, "ship")
        self.assertEqual([item.pk for item in transfers], [self.transfer.pk])

    def test_other_user_does_not_find_transfer_for_foreign_warehouse(self):
        transfers = find_active_transfers_by_barcode(self.other, self.product.barcode, "ship")
        self.assertEqual(transfers, [])

    def test_confirm_shipment_uses_database_quantity_and_prevents_repeat(self):
        confirm_shipment(self.sender, self.transfer.pk, self.product.barcode)
        self.assertEqual(Stock.objects.get(warehouse=self.source, product=self.product).quantity, 3)
        operation = OperationLog.objects.filter(transfer=self.transfer, action="Товар отгружен").latest("created_at")
        self.assertEqual(operation.metadata["barcode"], self.product.barcode)
        with self.assertRaises(ValidationError):
            confirm_shipment(self.sender, self.transfer.pk, self.product.barcode)

    def test_receiver_finds_and_confirms_in_transit_transfer(self):
        confirm_shipment(self.sender, self.transfer.pk, self.product.barcode)
        transfers = find_active_transfers_by_barcode(self.receiver, self.product.barcode, "receive")
        self.assertEqual([item.pk for item in transfers], [self.transfer.pk])
        confirm_receive(self.receiver, self.transfer.pk, self.product.barcode)
        self.assertEqual(Stock.objects.get(warehouse=self.destination, product=self.product).quantity, 2)
        self.transfer.refresh_from_db()
        self.assertEqual(self.transfer.status, Transfer.Status.COMPLETED)
